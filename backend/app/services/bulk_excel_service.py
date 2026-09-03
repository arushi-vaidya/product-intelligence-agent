"""
Bulk Excel / CSV intake.

This module does NOT change how a single product is investigated.
It reads a spreadsheet of products, runs each row through the
existing DFOO pipeline (``app.dfoo.orchestrator.DFOO``) exactly the
way a single manual investigation would, and then maps the resulting
``product_intelligence`` payload into the columns of the Delivery
Format spreadsheet so the whole batch can be downloaded as one
filled-in .xlsx file.
"""

import asyncio
import csv
import io
import re
import uuid
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Optional

from openpyxl import Workbook, load_workbook

from app.dfoo.orchestrator import DFOO
from app.dfoo.task_state import TaskStatus


# ==========================================
# INPUT COLUMNS
#
# These are the columns the uploaded sheet is
# expected to contain (matches the Unihack
# sample input dataset). Extra columns are
# ignored; missing ones are treated as blank.
# ==========================================

INPUT_COLUMNS = [
    "Mfg_Part_Num",
    "Part_Desc",
    "E1_Brand",
    "Unilog_Brand",
    "DIB_Brand",
    "Part_Manuf",
]


# ==========================================
# DELIVERY FORMAT COLUMNS
#
# Exact header row of the expected output
# spreadsheet. Order matters.
# ==========================================

DELIVERY_COLUMNS: list[str] = (
    [
        "MFR URL",
        "Ref URL 1",
        "Ref URL 2",
        "Ref URL 3",
        "Ref URL 4",
        "Ref URL 5",
        "PART_NUMBER",
        "Dept",
        "Class",
        "Fine",
        "SKU - MY_PART_NUMBER",
        "Mfg_Part_Num",
        "Part_Desc",
        "E1_Brand",
        "Unilog_Brand",
        "DIB_Brand",
        "Part_Manuf",
        "MANUFACTURER_NAME",
        "BRAND_NAME",
        "TRADE_NAME",
        "MANUFACTURER_PART_NUMBER",
        "ALTERNATE_PART_NUMBER",
        "Classpath",
        "MOBILE_DESC",
        "INVOICE_DESC",
        "SHORT_DESC",
        "LONG_DESC1",
        "RETAIL_DESC",
        "MARKETING_DESCRIPTION",
    ]
    + [f"ITEM_FEATURES_{i}" for i in range(1, 21)]
    + [
        "With",
        "Standard/Approvals",
        "Prop 65",
        "Application",
        "Includes",
        "Product Name",
    ]
    + [
        col
        for i in range(1, 51)
        for col in (
            f"ATTRIBUTE_LABEL {i}",
            f"ATTRIBUTE_VALUE {i}",
            f"ATTRIBUTE_UOM {i}",
        )
    ]
    + [
        "UPC",
        "EAN",
        "GTIN",
        "UNSPSC",
        "Warranty",
        "List Price",
        "Selling Qty",
        "Selling UOM",
        "Standard Packaging Information",
        "LENGTH",
        "LENGTH_UOM",
        "HEIGHT",
        "HEIGHT_UOM",
        "WIDTH",
        "WIDTH_UOM",
        "WEIGHT",
        "WEIGHT_UOM",
        "VOLUME",
        "VOLUME_UOM",
        "Product Image",
        "Alternate Image 1",
        "Alternate Image 2",
        "Alternate Image 3",
        "Alternate Image 4",
        "SDS",
        "SDS_1",
        "Warranty Information",
        "Catalog",
        "Specification Sheet",
        "Instruction/Installation Manual",
        "Service Manual",
        "Owners/User Manual",
        "Line Drawing",
        "MTR",
        "RoHS",
        "Full Engineering Drawing",
        "Energy Star Guide",
        "Technical Bulletin",
        "Submittal",
        "Compatibility Chart",
        "Size Chart",
        "Product Label/Insert",
        "Video Link",
        "Video Link 1",
        "Country Of Origin",
        "Discontinued",
        "Actual Image (Yes/No)",
    ]
)


# How many rows are allowed to be investigated
# concurrently. Each investigation makes several
# LLM / web-search calls, so this keeps us from
# hammering those APIs on a big upload.
# Each investigation already makes a large number of upstream LLM and web
# requests.  Running several complete pipelines at once regularly exhausts
# those providers' rate limits, even when a single investigation succeeds.
MAX_CONCURRENT_INVESTIGATIONS = 1

# Safety timeout per row so one stuck investigation
# can't hang the whole batch forever.
PER_ROW_TIMEOUT_SECONDS = 240
POLL_INTERVAL_SECONDS = 1.5


class BulkExcelError(Exception):
    pass


@dataclass
class BulkInvestigation:
    """In-memory state for a bulk job running in this API process."""

    id: str
    status: str = "pending"
    error: Optional[str] = None
    workbook_bytes: Optional[bytes] = None
    created_at: datetime = field(default_factory=datetime.utcnow)


class BulkInvestigationRepository:
    def __init__(self) -> None:
        self.jobs: dict[str, BulkInvestigation] = {}

    def create(self) -> BulkInvestigation:
        job = BulkInvestigation(id=str(uuid.uuid4()))
        self.jobs[job.id] = job
        return job

    def get(self, job_id: str) -> Optional[BulkInvestigation]:
        return self.jobs.get(job_id)


bulk_investigation_repository = BulkInvestigationRepository()


# ==========================================
# READING THE UPLOADED SHEET
# ==========================================

def read_input_rows(
    filename: str,
    content: bytes,
) -> list[dict[str, str]]:
    """
    Parse an uploaded .xlsx/.xlsm/.csv file into a
    list of row dicts keyed by the input columns.
    """

    lower_name = (filename or "").lower()

    try:
        if lower_name.endswith(".csv"):
            rows = _read_csv_rows(content)
        elif lower_name.endswith((".xlsx", ".xlsm")):
            rows = _read_xlsx_rows(content)
        else:
            raise BulkExcelError(
                "Unsupported file type. Please upload a "
                ".xlsx, .xlsm, or .csv file."
            )
    except BulkExcelError:
        raise
    except (
        OSError,
        UnicodeDecodeError,
        ValueError,
        csv.Error,
        zipfile.BadZipFile,
    ) as error:
        raise BulkExcelError(
            "The uploaded spreadsheet could not be read. "
            "Please upload a valid .xlsx, .xlsm, or UTF-8 CSV file."
        ) from error

    if not rows:
        raise BulkExcelError(
            "No product rows were found in the "
            "uploaded file."
        )

    return rows


def _read_csv_rows(
    content: bytes,
) -> list[dict[str, str]]:

    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    return [
        _normalize_row(row)
        for row in reader
        if any((value or "").strip() for value in row.values())
    ]


def _read_xlsx_rows(
    content: bytes,
) -> list[dict[str, str]]:

    workbook = load_workbook(
        io.BytesIO(content),
        data_only=True,
        read_only=True,
    )

    sheet = workbook.active

    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    header = [
        str(cell).strip() if cell is not None else ""
        for cell in header
    ]

    rows: list[dict[str, str]] = []

    for raw_row in rows_iter:

        if raw_row is None or all(
            cell is None for cell in raw_row
        ):
            continue

        row = {
            header[i]: (
                "" if raw_row[i] is None else str(raw_row[i])
            )
            for i in range(len(header))
            if i < len(raw_row)
        }

        rows.append(_normalize_row(row))

    return rows


def _normalize_row(
    row: dict[str, Any],
) -> dict[str, str]:

    normalized: dict[str, str] = {}

    for column in INPUT_COLUMNS:
        value = row.get(column, "")
        normalized[column] = (
            "" if value is None else str(value).strip()
        )

    return normalized


# ==========================================
# RUNNING THE (UNCHANGED) INVESTIGATION
# PIPELINE FOR EACH ROW
# ==========================================

SUPPLIER_CODE_SUFFIX = re.compile(r"\s*\([^()]*\)\s*$")
PLACEHOLDER_BRANDS = {
    "",
    "na",
    "none",
    "unknown",
    "unbranded",
    "nounilogbrand",
}
MPN_BRAND_PREFIXES = {
    "3M": "3M",
}


def _clean_manufacturer_name(name: str) -> str:
    """
    Some supplier fields append an internal distributor
    code, for example "Freud Inc (2435)". Strip it before
    using the value as an investigation input.
    """

    return SUPPLIER_CODE_SUFFIX.sub("", name).strip()


def _is_placeholder_brand(value: str) -> bool:
    """Treat display placeholders such as '-- Unbranded --' as blank."""

    normalized = re.sub(r"[^a-z0-9]", "", value.lower())
    return (
        normalized in PLACEHOLDER_BRANDS
        or (
            normalized.startswith("no")
            and normalized.endswith("brand")
        )
    )


def _infer_brand_from_mpn(mpn: str) -> str:
    normalized_mpn = re.sub(r"[^A-Za-z0-9]", "", mpn).upper()

    for prefix, brand in MPN_BRAND_PREFIXES.items():
        if normalized_mpn.startswith(prefix):
            return brand

    return ""


def _build_investigation_input(
    row: dict[str, str],
) -> Optional[dict[str, str]]:

    mpn = (row.get("Mfg_Part_Num") or "").strip()

    # Part_Manuf is often a distributor, so use a usable brand column
    # first. If those are all placeholders, infer a known brand prefix
    # from the MPN before falling back to the supplier value.
    raw_manufacturer = next(
        (
            value.strip()
            for column in (
                "E1_Brand",
                "Unilog_Brand",
                "DIB_Brand",
            )
            if (value := row.get(column, "")).strip()
            and not _is_placeholder_brand(value)
        ),
        "",
    )

    raw_manufacturer = (
        raw_manufacturer
        or _infer_brand_from_mpn(mpn)
        or (row.get("Part_Manuf") or "").strip()
    )

    manufacturer = _clean_manufacturer_name(raw_manufacturer)

    if not manufacturer or not mpn:
        return None

    return {
        "manufacturer": manufacturer,
        "mpn": mpn,
    }


async def _run_single_row(
    dfoo: DFOO,
    row: dict[str, str],
    semaphore: asyncio.Semaphore,
) -> dict[str, Any]:

    investigation_input = _build_investigation_input(row)

    if investigation_input is None:
        return {
            "row": row,
            "product_intelligence": None,
            "error": (
                "Missing manufacturer/brand or "
                "Mfg_Part_Num - skipped."
            ),
        }

    try:
        async with semaphore:

            # This is the exact same call the single-product
            # "New Investigation" flow makes. Nothing about
            # the pipeline itself changes for bulk uploads.
            investigation = await dfoo.start_investigation(
                investigation_input
            )

            elapsed = 0.0

            while elapsed < PER_ROW_TIMEOUT_SECONDS:

                current = dfoo.investigation_repository.get(
                    investigation.id
                )

                if current is None:
                    break

                if current.status == TaskStatus.DONE:
                    product_intelligence = (
                        current.result.get(
                            "product_intelligence",
                            current.result,
                        )
                        if current.result
                        else None
                    )

                    return {
                        "row": row,
                        "product_intelligence": product_intelligence,
                        "error": None,
                    }

                if current.status == TaskStatus.FAILED:
                    return {
                        "row": row,
                        "product_intelligence": None,
                        "error": "Investigation failed.",
                    }

                await asyncio.sleep(POLL_INTERVAL_SECONDS)
                elapsed += POLL_INTERVAL_SECONDS

            return {
                "row": row,
                "product_intelligence": None,
                "error": "Investigation timed out.",
            }
    except Exception as error:
        # One bad upstream response must not discard an entire uploaded sheet.
        return {
            "row": row,
            "product_intelligence": None,
            "error": f"Investigation error: {error}",
        }


async def process_rows(
    dfoo: DFOO,
    rows: list[dict[str, str]],
) -> list[dict[str, Any]]:

    semaphore = asyncio.Semaphore(
        MAX_CONCURRENT_INVESTIGATIONS
    )

    tasks = [
        _run_single_row(dfoo, row, semaphore)
        for row in rows
    ]

    return await asyncio.gather(*tasks)


# ==========================================
# MAPPING RESULTS -> DELIVERY FORMAT
# ==========================================

def _text(value: Any) -> str:

    if value is None:
        return ""

    if isinstance(value, (dict, list)):
        return ""

    return str(value)


def _extract_texts(
    entries: Any,
    key: str = "text",
) -> list[str]:
    """
    Enrichment lists (features, applications) come back
    as either [{"text": ..., "supported_by": [...]}] or
    occasionally plain strings, depending on the LLM
    output - handle both.
    """

    if not isinstance(entries, list):
        return []

    texts = []

    for entry in entries:

        if isinstance(entry, dict):
            text = entry.get(key)

            if text:
                texts.append(str(text))

        elif isinstance(entry, str):
            texts.append(entry)

    return texts


def _is_valid_identity(value: str) -> bool:
    return bool(value.strip()) and not _is_placeholder_brand(value)


def _brand_from_title(title: str) -> str:
    """Use the leading product-brand token only as a fallback."""

    first_word = title.strip().split(maxsplit=1)

    if not first_word:
        return ""

    return first_word[0].strip("™®:,-")


def _is_abrasive_product(title: str, description: str) -> bool:
    text = f"{title} {description}".lower()
    keywords = (
        "abrasive",
        "sanding",
        "sandpaper",
        "grit",
        "film disc",
        "sanding belt",
    )
    return any(keyword in text for keyword in keywords)


def map_to_delivery_row(
    result: dict[str, Any],
) -> dict[str, str]:

    row = result.get("row", {})
    pi = result.get("product_intelligence") or {}
    error = result.get("error")

    delivery: dict[str, str] = {
        column: "" for column in DELIVERY_COLUMNS
    }

    # -----------------------------------
    # Passthrough input columns
    # -----------------------------------

    for column in INPUT_COLUMNS:
        delivery[column] = row.get(column, "")

    delivery["PART_NUMBER"] = row.get("Mfg_Part_Num", "")

    if not pi:
        # No successful investigation for this row -
        # leave the enrichment columns blank but keep
        # a note so the row can be spotted and re-run.
        delivery["INVOICE_DESC"] = (
            error or "No data returned."
        )
        return delivery

    # -----------------------------------
    # Manufacturer / brand
    # -----------------------------------

    resolved_input = _build_investigation_input(row) or {}
    manufacturer = _text(pi.get("manufacturer")).strip()

    if not _is_valid_identity(manufacturer):
        manufacturer = resolved_input.get("manufacturer", "")

    mpn = _text(pi.get("mpn")) or row.get("Mfg_Part_Num", "")

    delivery["MANUFACTURER_NAME"] = manufacturer
    delivery["MANUFACTURER_PART_NUMBER"] = mpn

    delivery["Classpath"] = _text(
        pi.get("product_category")
    )

    # -----------------------------------
    # Enrichment -> descriptions / features
    # -----------------------------------

    enrichment = pi.get("enrichment") or {}

    title = _text(enrichment.get("title"))
    product_name = title or row.get(
        "Part_Desc", ""
    )

    delivery["Product Name"] = product_name

    e1_brand = row.get("E1_Brand", "").strip()
    delivery["BRAND_NAME"] = (
        e1_brand
        if _is_valid_identity(e1_brand)
        else _brand_from_title(product_name) or manufacturer
    )

    short_description = enrichment.get(
        "short_description"
    )

    short_text = (
        short_description.get("text")
        if isinstance(short_description, dict)
        else short_description
    )

    delivery["SHORT_DESC"] = _text(short_text)
    delivery["MARKETING_DESCRIPTION"] = _text(short_text)
    delivery["MOBILE_DESC"] = _text(short_text)
    delivery["RETAIL_DESC"] = _text(short_text)

    features = _extract_texts(enrichment.get("features"))

    for index, feature_text in enumerate(features[:20], start=1):
        delivery[f"ITEM_FEATURES_{index}"] = feature_text

    if features:
        delivery["LONG_DESC1"] = " ".join(features)

    applications = _extract_texts(
        enrichment.get("applications")
    )

    delivery["Application"] = "; ".join(applications)

    # -----------------------------------
    # Family specifications -> attributes
    # -----------------------------------

    family_specifications = (
        pi.get("family_specifications") or {}
    )

    electrical_only_fields = {
        "rated_current",
        "poles",
        "trip_curve",
        "frequency",
        "breaking_capacity",
    }

    if _is_abrasive_product(product_name, _text(short_text)):
        family_specifications = {
            field: specification
            for field, specification in family_specifications.items()
            if field not in electrical_only_fields
        }

    for index, (field_name, spec) in enumerate(
        list(family_specifications.items())[:50],
        start=1,
    ):
        value = (
            spec.get("value")
            if isinstance(spec, dict)
            else spec
        )

        unit = (
            spec.get("unit")
            if isinstance(spec, dict)
            else None
        )

        delivery[f"ATTRIBUTE_LABEL {index}"] = field_name
        delivery[f"ATTRIBUTE_VALUE {index}"] = _text(value)
        delivery[f"ATTRIBUTE_UOM {index}"] = _text(unit)

    # -----------------------------------
    # Sources -> reference URLs
    # -----------------------------------

    sources = pi.get("sources") or []
    urls = [
        source.get("url")
        for source in sources
        if isinstance(source, dict) and source.get("url")
    ]

    manufacturer_urls = [
        source.get("url")
        for source in sources
        if isinstance(source, dict)
        and source.get("source_type")
        in {"manufacturer_page", "manufacturer_datasheet"}
        and source.get("url")
    ]

    if manufacturer_urls:
        delivery["MFR URL"] = manufacturer_urls[0]

    for index, url in enumerate(urls[:5], start=1):
        delivery[f"Ref URL {index}"] = url

    # -----------------------------------
    # Commerce readiness note
    # -----------------------------------

    commerce_readiness = (
        pi.get("commerce_readiness") or {}
    )

    if commerce_readiness.get("status") == (
        "review_required"
    ):
        delivery["INVOICE_DESC"] = (
            "Flagged for human review by the "
            "pipeline."
        )

    return delivery


# ==========================================
# WRITING THE OUTPUT WORKBOOK
# ==========================================

def build_output_workbook(
    delivery_rows: list[dict[str, str]],
) -> bytes:

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Delivery Format"

    sheet.append(DELIVERY_COLUMNS)

    for delivery_row in delivery_rows:
        sheet.append(
            [
                delivery_row.get(column, "")
                for column in DELIVERY_COLUMNS
            ]
        )

    buffer = io.BytesIO()
    workbook.save(buffer)

    return buffer.getvalue()


async def run_bulk_excel(
    dfoo: DFOO,
    filename: str,
    content: bytes,
) -> bytes:
    """
    End-to-end: parse the uploaded sheet, run every row
    through the existing DFOO pipeline, and return the
    filled Delivery Format workbook as .xlsx bytes.
    """

    rows = read_input_rows(filename, content)

    results = await process_rows(dfoo, rows)

    delivery_rows = [
        map_to_delivery_row(result) for result in results
    ]

    return build_output_workbook(delivery_rows)


def start_bulk_excel_investigation(
    dfoo: DFOO,
    filename: str,
    content: bytes,
) -> BulkInvestigation:
    """Start a bulk job without holding the upload HTTP request open."""

    # Validate input before acknowledging the upload, so format errors are
    # returned to the user immediately rather than becoming opaque job errors.
    read_input_rows(filename, content)

    job = bulk_investigation_repository.create()

    async def run_job() -> None:
        job.status = "running"
        try:
            job.workbook_bytes = await run_bulk_excel(
                dfoo=dfoo,
                filename=filename,
                content=content,
            )
            job.status = "done"
        except Exception as error:
            job.error = str(error) or "Bulk investigation failed."
            job.status = "failed"

    asyncio.create_task(run_job())
    return job
